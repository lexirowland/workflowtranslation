import difflib
from difflib import SequenceMatcher
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

wb = load_workbook(filename = 'similarityassessmentA.xlsx')
ws = wb['CompetencyQ5']

r = 10
col = 9
resultr = 10
resultcol = col + 2

count = 0
while count < 19:

    #1
    #data input type
    str1 = ws.cell(row = r, column = col).value
    str2 = ws.cell(row = r, column = col + 1).value

    seq = SequenceMatcher(a=str1, b=str2)

    print("sequence ratio is: {}".format(seq.ratio()))

    ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    wb.save('similarityassessmentA.xlsx')

    r += 1
    resultr += 1
    #data output type
    str3 = ws.cell(row = r, column = col).value
    str4 = ws.cell(row = r, column = col + 1).value

    seq = SequenceMatcher(a=str3, b=str4)

    print("sequence ratio is: {}".format(seq.ratio()))

    ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    wb.save('similarityassessmentA.xlsx')

    r += 1
    resultr += 1
    # #tool label
    str5 = ws.cell(row = r, column = col).value
    str6 = ws.cell(row = r, column = col + 1).value

    seq = SequenceMatcher(a=str5, b=str6)

    print("sequence ratio is: {}".format(seq.ratio()))

    ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    wb.save('similarityassessmentA.xlsx')

    #2
    r += 2
    resultr +=2
    # #data input type
    str7 = ws.cell(row = r, column = col).value
    str8 = ws.cell(row = r, column = col + 1).value

    seq = SequenceMatcher(a=str7, b=str8)

    ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    wb.save('similarityassessmentA.xlsx')

    r += 1
    resultr +=1
    # #data output type
    str9 = ws.cell(row = r, column = col).value
    str10 = ws.cell(row = r, column = col + 1).value

    seq = SequenceMatcher(a=str9, b=str10)

    ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    wb.save('similarityassessmentA.xlsx')

    r += 1
    resultr += 1
    # #tool label
    str11 = ws.cell(row = r, column = col).value
    str12 = ws.cell(row = r, column = col + 1).value

    seq = SequenceMatcher(a=str11, b=str12)

    ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    wb.save('similarityassessmentA.xlsx')

    # #3
    # r += 2
    # resultr +=2
    # # #data input type
    # str11 = ws.cell(row = r, column = col).value
    # str12 = ws.cell(row = r, column = col + 1).value

    # seq = SequenceMatcher(a=str11, b=str12)

    # ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    # wb.save('similarityassessmentA.xlsx')

    # r += 1
    # resultr += 1
    # # #data output type
    # str13 = ws.cell(row = r, column = col).value
    # str14 = ws.cell(row = r, column = col + 1).value

    # seq = SequenceMatcher(a=str13, b=str14)

    # ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    # wb.save('similarityassessmentA.xlsx')

    # r += 1
    # resultr += 1
    # # #tool label
    # str15 = ws.cell(row = r, column = col).value
    # str16 = ws.cell(row = r, column = col + 1).value

    # seq = SequenceMatcher(a=str15, b=str16)

    # ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    # wb.save('similarityassessmentA.xlsx')

    # #4
    # r += 2
    # resultr +=2
    # # #data input type
    # str17 = ws.cell(row = r, column = col).value
    # str18 = ws.cell(row = r, column = col + 1).value

    # seq = SequenceMatcher(a=str17, b=str18)

    # ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    # wb.save('similarityassessmentA.xlsx')

    # r += 1
    # resultr += 1
    # # #data output type
    # str19 = ws.cell(row = r, column = col).value
    # str20 = ws.cell(row = r, column = col + 1).value

    # seq = SequenceMatcher(a=str19, b=str20)

    # ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    # wb.save('similarityassessmentA.xlsx')

    # r += 1
    # resultr += 1
    # # #tool label
    # str21 = ws.cell(row = r, column = col).value
    # str22 = ws.cell(row = r, column = col + 1).value

    # seq = SequenceMatcher(a=str21, b=str22)

    # ws.cell(row = resultr, column = resultcol).value = seq.ratio()
    # wb.save('similarityassessmentA.xlsx')

    r += 6
    resultr += 6

    count += 1